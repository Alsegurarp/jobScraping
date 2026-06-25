#!/usr/bin/env python3
import argparse
import html
import json
import re
import textwrap
import urllib.parse
import urllib.request
from datetime import date, datetime, timedelta
from email.utils import parsedate_to_datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with 'python -m pip install openpyxl' "
        "or run this script with Codex bundled Python."
    ) from exc


INPUT_SHEET = "vacantes"
OUTPUT_SHEETS = [
    "vacantes_detectadas",
    "preseleccionadas",
    "descartadas",
    "aplicadas",
    "empresas_investigadas",
]

INPUT_COLUMNS = [
    "nombre_de_la_vacante",
    "empresa",
    "portal",
    "industria",
    "ubicacion",
    "modalidad",
    "salario",
    "horas_semana",
    "seniority",
    "idioma",
    "fecha_publicacion",
    "url",
    "email_contacto",
    "descripcion",
    "url_empresa",
]

RESULT_COLUMNS = [
    "prioridad",
    "score",
    "estado",
    "nombre_de_la_vacante",
    "empresa",
    "portal",
    "industria",
    "ubicacion",
    "modalidad",
    "salario",
    "horas_semana",
    "seniority",
    "idioma",
    "url",
    "email_contacto",
    "documento_que_se_manda",
    "carta_de_interes_al_rol",
    "mensaje_corto_reclutador",
    "razon_menos_250",
    "matched_skills",
    "flags",
]

RESEARCH_COLUMNS = [
    "empresa",
    "resumen",
    "fuentes",
    "fecha_investigacion",
]

STOPWORDS = {
    "a", "al", "and", "con", "de", "del", "el", "en", "for", "la", "las",
    "los", "of", "para", "por", "the", "to", "un", "una", "y", "or", "o",
}


def words(text):
    return {
        word
        for word in re.findall(r"[a-z0-9+#.]+", (text or "").lower())
        if len(word) > 2 and word not in STOPWORDS
    }


def clean_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def slug(value):
    text = clean_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-") or "item"


def load_profile(path):
    profile = json.loads(Path(path).read_text(encoding="utf-8"))
    skill_terms = words(" ".join(profile.get("skills", [])))
    interest_terms = words(" ".join(profile.get("interest_keywords", [])))
    profile["skill_terms"] = skill_terms
    profile["interest_terms"] = interest_terms
    return profile


def read_sheet(path):
    workbook = load_workbook(path)
    if INPUT_SHEET not in workbook.sheetnames:
        raise SystemExit(f"Input workbook must include a '{INPUT_SHEET}' sheet.")
    sheet = workbook[INPUT_SHEET]
    headers = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(values):
            continue
        row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        rows.append(row)
    return rows


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    try:
        return parsedate_to_datetime(text).date()
    except Exception:
        return None


def parse_money_mxn(value):
    text = clean_text(value).lower()
    if not text:
        return None
    amounts = [float(number.replace(",", "")) for number in re.findall(r"\d[\d,]*(?:\.\d+)?", text)]
    if not amounts:
        return None
    amount = max(amounts)
    if "usd" in text or "dolar" in text or "dollar" in text:
        return amount * 18.5
    if amount < 1000 and ("hora" in text or "hour" in text):
        return amount * 40 * 4.33
    return amount


def is_remote(row):
    text = " ".join(clean_text(row.get(key)) for key in ("modalidad", "ubicacion", "descripcion")).lower()
    return any(term in text for term in ("remoto", "remote", "home office", "work from home"))


def is_hybrid(row):
    text = " ".join(clean_text(row.get(key)) for key in ("modalidad", "ubicacion", "descripcion")).lower()
    return "hibrid" in text or "hybrid" in text


def in_cdmx(row):
    text = " ".join(clean_text(row.get(key)) for key in ("ubicacion", "descripcion")).lower()
    return any(term in text for term in ("cdmx", "ciudad de mexico", "ciudad de méxico", "mexico city"))


def mentions_project_work(text):
    terms = [
        "por proyecto", "project based", "project-based", "proyecto temporal",
        "contrato por proyecto", "temporary project",
    ]
    return any(term in text for term in terms)


def quality_flags(profile, row, matched_skills):
    text = " ".join(clean_text(value) for value in row.values()).lower()
    flags = []
    industry = clean_text(row.get("industria")).lower()

    allowed_industries = [item.lower() for item in profile.get("allowed_industries", [])]
    blocked_industries = [item.lower() for item in profile.get("blocked_industries", [])]
    if allowed_industries and industry and not any(item in industry for item in allowed_industries):
        flags.append("industria_fuera_de_foco")
    if any(item in industry or item in text for item in blocked_industries):
        flags.append("industria_bloqueada")

    if mentions_project_work(text):
        flags.append("trabajo_por_proyecto")
    if any(term in text for term in ("guardia", "nocturno", "night shift", "weekend", "fin de semana", "24/7", "alta disponibilidad")):
        flags.append("horario_descartable")
    if "freelance" in text and any(term in text for term in ("exclusividad", "exclusive", "full dedication")):
        flags.append("freelance_con_exclusividad")

    hours = clean_text(row.get("horas_semana"))
    hour_numbers = [int(number) for number in re.findall(r"\d+", hours)]
    if hour_numbers and max(hour_numbers) > profile.get("max_hours_per_week", 40):
        flags.append("mas_de_40_horas")

    if not is_remote(row) and not is_hybrid(row) and not in_cdmx(row):
        flags.append("presencial_fuera_cdmx")
    if is_hybrid(row) and not in_cdmx(row):
        flags.append("hibrido_fuera_cdmx")

    seniority = " ".join([clean_text(row.get("seniority")), clean_text(row.get("nombre_de_la_vacante"))]).lower()
    if any(term in seniority for term in ("senior", "sr.", "lead", "principal", "staff")):
        flags.append("seniority_alto")

    salary = parse_money_mxn(row.get("salario"))
    if salary is not None and salary < profile.get("minimum_salary_mxn", 20000):
        flags.append("salario_bajo")

    if len(matched_skills) < profile.get("minimum_skill_matches", 5):
        flags.append("menos_de_5_skills")

    spam_terms = ("multinivel", "sin experiencia gana", "pago inicial", "curso obligatorio", "deposito", "inversion inicial")
    if any(term in text for term in spam_terms):
        flags.append("posible_spam")

    return flags


def score_job(profile, row):
    text = " ".join(clean_text(value) for value in row.values())
    job_terms = words(text)
    matched_skills = sorted(profile["skill_terms"] & job_terms)
    matched_interests = sorted(profile["interest_terms"] & job_terms)
    flags = quality_flags(profile, row, matched_skills)

    score = 0
    industry = clean_text(row.get("industria")).lower()
    if any(item in industry for item in [x.lower() for x in profile.get("allowed_industries", [])]):
        score += 22
    if parse_money_mxn(row.get("salario")) and parse_money_mxn(row.get("salario")) >= profile.get("minimum_salary_mxn", 20000):
        score += 18
    elif not clean_text(row.get("salario")):
        score += 8
    if not any(flag in flags for flag in ("mas_de_40_horas", "horario_descartable")):
        score += 14
    if is_remote(row):
        score += 14
    elif is_hybrid(row) and in_cdmx(row):
        score += 8
    elif in_cdmx(row):
        score += 5
    if any(term in " ".join([clean_text(row.get("seniority")), clean_text(row.get("nombre_de_la_vacante"))]).lower() for term in ("junior", "jr", "trainee")):
        score += 12
    score += min(len(matched_skills) * 3, 15)
    score += min(len(matched_interests) * 2, 5)

    hard_flags = {
        "industria_bloqueada", "trabajo_por_proyecto", "horario_descartable",
        "mas_de_40_horas", "presencial_fuera_cdmx", "hibrido_fuera_cdmx",
        "seniority_alto", "salario_bajo", "menos_de_5_skills", "posible_spam",
    }
    if any(flag in hard_flags for flag in flags):
        status = "descartada"
        score = min(score, 59)
    else:
        status = "preseleccionada"

    return min(score, 100), status, matched_skills, matched_interests, flags


def fetch_url_text(url, timeout=12):
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "Mozilla/5.0 BotJobs local research"},
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        content_type = response.headers.get("content-type", "")
        raw = response.read(300000)
    encoding = "utf-8"
    match = re.search(r"charset=([\w-]+)", content_type)
    if match:
        encoding = match.group(1)
    text = raw.decode(encoding, errors="ignore")
    text = re.sub(r"<(script|style).*?</\1>", " ", text, flags=re.I | re.S)
    text = re.sub(r"<[^>]+>", " ", text)
    return html.unescape(re.sub(r"\s+", " ", text)).strip()


def search_company(company):
    query = urllib.parse.quote_plus(f"{company} empresa tecnologia software")
    url = f"https://duckduckgo.com/html/?q={query}"
    return fetch_url_text(url)


def research_company(row, enabled):
    company = clean_text(row.get("empresa"))
    if not company:
        return {"empresa": "", "resumen": "", "fuentes": "", "fecha_investigacion": ""}
    if not enabled:
        return {"empresa": company, "resumen": "Investigacion web no ejecutada.", "fuentes": "", "fecha_investigacion": ""}

    sources = []
    texts = []
    company_url = clean_text(row.get("url_empresa"))
    if company_url:
        try:
            texts.append(fetch_url_text(company_url))
            sources.append(company_url)
        except Exception as exc:
            texts.append(f"No se pudo leer {company_url}: {exc}")
    if not texts:
        try:
            texts.append(search_company(company))
            sources.append("DuckDuckGo HTML search")
        except Exception as exc:
            texts.append(f"No se pudo investigar la empresa: {exc}")

    joined = " ".join(texts)
    summary = clean_text(joined[:650])
    return {
        "empresa": company,
        "resumen": summary,
        "fuentes": "; ".join(sources),
        "fecha_investigacion": date.today().isoformat(),
    }


def short_reason(status, score, flags, matched_skills):
    if status == "preseleccionada":
        reason = f"Buena afinidad ({score}/100): coincide con {len(matched_skills)} skills clave y respeta filtros base."
    else:
        reason = f"Descartada ({score}/100): {', '.join(flags[:3]) or 'baja afinidad con el perfil'}."
    return reason[:247] + "..." if len(reason) > 250 else reason


def detect_language(row):
    explicit = clean_text(row.get("idioma")).lower()
    if explicit.startswith("en") or "ingles" in explicit or "english" in explicit:
        return "en"
    text = " ".join(clean_text(row.get(key)) for key in ("nombre_de_la_vacante", "descripcion")).lower()
    english_hits = sum(term in text for term in ("developer", "engineer", "remote", "requirements", "responsibilities", "english"))
    spanish_hits = sum(term in text for term in ("desarrollador", "remoto", "requisitos", "responsabilidades", "experiencia"))
    return "en" if english_hits > spanish_hits else "es"


def cover_letter(profile, row, matched_skills, research):
    language = detect_language(row)
    name = profile.get("name", "Rene Alexis Segura Perez")
    role = clean_text(row.get("nombre_de_la_vacante")) or "the role"
    company = clean_text(row.get("empresa")) or "your company"
    skills = ", ".join(matched_skills[:8]) or ", ".join(profile.get("skills", [])[:6])
    github = profile.get("github", "")
    portfolio = profile.get("portfolio", "")
    availability = profile.get("availability", "")
    no_salary = not clean_text(row.get("salario"))
    research_summary = clean_text(research.get("resumen"))[:450]

    if language == "en":
        salary_line = (
            "Since the salary range is not published, I would be interested if the role is above "
            "MXN $22,000 per month or its USD equivalent.\n\n"
            if no_salary else ""
        )
        return f"""# Cover letter - {role} - {company}

Hello {company} team,

My name is {name}. I am a Full Stack Developer focused on React, TypeScript, Node.js, NestJS, CI/CD and Azure DevOps. I am interested in the {role} position because it matches my experience with {skills}.

Based on the role and what I found about {company}, I can contribute by building maintainable web applications, improving delivery workflows, debugging production issues and supporting reliable deployments with Docker, Linux, Nginx and GitHub Actions.

Company context considered:
{research_summary or "No public company research was available in this run."}

{salary_line}My availability is {availability}. You can review my work here:
- GitHub: {github}
- Portfolio: {portfolio}

Best regards,
{name}
"""

    salary_line = (
        "Como la oferta no publica rango salarial, me interesa avanzar si la posicion supera los "
        "$22,000 MXN mensuales o su equivalente en dolares.\n\n"
        if no_salary else ""
    )
    return f"""# Carta de interes - {role} - {company}

Hola equipo de {company},

Mi nombre es {name}. Soy Full Stack Developer con enfoque en React, TypeScript, Node.js, NestJS, CI/CD y Azure DevOps. Me interesa la vacante de {role} porque conecta directamente con mi experiencia en {skills}.

Por el tipo de rol y lo investigado sobre {company}, puedo aportar valor construyendo aplicaciones web mantenibles, mejorando flujos de entrega, resolviendo incidencias en produccion y apoyando despliegues confiables con Docker, Linux, Nginx y GitHub Actions.

Contexto de empresa considerado:
{research_summary or "No se ejecuto investigacion publica de empresa en esta corrida."}

{salary_line}Mi disponibilidad es {availability}. Puedes revisar mi trabajo aqui:
- GitHub: {github}
- Portafolio: {portfolio}

Saludos,
{name}
"""


def recruiter_message(profile, row, matched_skills):
    language = detect_language(row)
    name = profile.get("name", "Rene Alexis Segura Perez")
    role = clean_text(row.get("nombre_de_la_vacante")) or "the role"
    skills = ", ".join(matched_skills[:5]) or "React, TypeScript, Node.js, NestJS, CI/CD"
    if language == "en":
        return (
            f"Hi, I am {name}, a Junior Full Stack Developer with hands-on experience in {skills}. "
            f"I am interested in {role} and can support product delivery, debugging and reliable deployments."
        )
    return (
        f"Hola, soy {name}, Junior Full Stack Developer con experiencia practica en {skills}. "
        f"Me interesa la vacante de {role} y puedo aportar en desarrollo, debugging y despliegues confiables."
    )


def write_rows(sheet, headers, rows):
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    style_sheet(sheet)


def style_sheet(sheet):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = get_column_letter(column[0].column)
        max_len = max(len(clean_text(cell.value)) for cell in column)
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)


def write_output(path, detected, shortlisted, discarded, applied, research_rows):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in OUTPUT_SHEETS:
        workbook.create_sheet(name)
    write_rows(workbook["vacantes_detectadas"], RESULT_COLUMNS, detected)
    write_rows(workbook["preseleccionadas"], RESULT_COLUMNS, shortlisted)
    write_rows(workbook["descartadas"], RESULT_COLUMNS, discarded)
    write_rows(workbook["aplicadas"], RESULT_COLUMNS, applied)
    write_rows(workbook["empresas_investigadas"], RESEARCH_COLUMNS, research_rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def create_template(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = INPUT_SHEET
    sheet.append(INPUT_COLUMNS)
    sheet.append([
        "Junior Full Stack Developer",
        "Acme SaaS",
        "LinkedIn",
        "SaaS",
        "Remoto Mexico",
        "remoto",
        "25000 MXN",
        40,
        "Junior",
        "es",
        date.today().isoformat(),
        "https://example.com/job",
        "recruiter@example.com",
        "Buscamos developer con React, TypeScript, Node.js, APIs REST, PostgreSQL, Docker y CI/CD.",
        "https://example.com",
    ])
    style_sheet(sheet)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def run(profile_path, jobs_path, output_dir, research_enabled=False):
    profile = load_profile(profile_path)
    rows = read_sheet(jobs_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    letters_dir = output_dir / "cartas"
    letters_dir.mkdir(parents=True, exist_ok=True)

    detected = []
    shortlisted = []
    discarded = []
    research_by_company = {}
    cv_path = profile.get("cv_file", "Rene_Alexis_Segura_CV.pdf")

    cutoff = date.today() - timedelta(days=profile.get("max_post_age_days", 14))

    for row in rows:
        posted = parse_date(row.get("fecha_publicacion"))
        if posted and posted < cutoff:
            row = {**row, "descripcion": f"{clean_text(row.get('descripcion'))} vacante_mayor_a_2_semanas"}

        score, status, matched_skills, _matched_interests, flags = score_job(profile, row)
        company = clean_text(row.get("empresa"))
        if company not in research_by_company:
            research_by_company[company] = research_company(row, research_enabled)
        research = research_by_company[company]

        letter_path = letters_dir / f"{slug(company)}-{slug(row.get('nombre_de_la_vacante'))}.md"
        message = recruiter_message(profile, row, matched_skills)
        letter_path.write_text(cover_letter(profile, row, matched_skills, research), encoding="utf-8")

        result = {
            "prioridad": "",
            "score": score,
            "estado": status,
            "nombre_de_la_vacante": clean_text(row.get("nombre_de_la_vacante")),
            "empresa": company,
            "portal": clean_text(row.get("portal")),
            "industria": clean_text(row.get("industria")),
            "ubicacion": clean_text(row.get("ubicacion")),
            "modalidad": clean_text(row.get("modalidad")),
            "salario": clean_text(row.get("salario")),
            "horas_semana": clean_text(row.get("horas_semana")),
            "seniority": clean_text(row.get("seniority")),
            "idioma": detect_language(row),
            "url": clean_text(row.get("url")),
            "email_contacto": clean_text(row.get("email_contacto")),
            "documento_que_se_manda": cv_path,
            "carta_de_interes_al_rol": str(letter_path),
            "mensaje_corto_reclutador": message,
            "razon_menos_250": short_reason(status, score, flags, matched_skills),
            "matched_skills": ", ".join(matched_skills),
            "flags": ", ".join(flags),
        }
        detected.append(result)
        if status == "preseleccionada":
            shortlisted.append(result)
        else:
            discarded.append(result)

    detected.sort(key=lambda item: item["score"], reverse=True)
    shortlisted.sort(key=lambda item: item["score"], reverse=True)
    discarded.sort(key=lambda item: item["score"], reverse=True)
    for index, item in enumerate(shortlisted, 1):
        item["prioridad"] = index
    for index, item in enumerate(detected, 1):
        item["prioridad"] = item["prioridad"] or index

    output_path = output_dir / "botjobs_resultados.xlsx"
    write_output(output_path, detected, shortlisted, discarded, [], list(research_by_company.values()))
    return output_path


def demo():
    template = Path("vacantes.template.xlsx")
    if not template.exists():
        create_template(template)
    output_path = run(Path("profile.example.json"), template, Path("output"), research_enabled=False)
    print(f"Demo ok: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Bot local para rankear vacantes tech y generar cartas.")
    parser.add_argument("--profile", default="profile.example.json")
    parser.add_argument("--jobs", default="vacantes.template.xlsx")
    parser.add_argument("--out", default="output")
    parser.add_argument("--research", action="store_true", help="Investiga empresas por internet.")
    parser.add_argument("--create-template", action="store_true", help="Crea una plantilla .xlsx de vacantes.")
    parser.add_argument("--demo", action="store_true")
    args = parser.parse_args()

    if args.create_template:
        create_template(Path(args.jobs))
        print(f"Plantilla creada: {args.jobs}")
        return
    if args.demo:
        demo()
        return

    output_path = run(Path(args.profile), Path(args.jobs), Path(args.out), args.research)
    print(f"Listo: {output_path}")


if __name__ == "__main__":
    main()
