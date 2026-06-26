from datetime import date, datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with 'python -m pip install openpyxl' "
        "or run this script with Codex bundled Python."
    ) from exc

from .schema import (
    INPUT_COLUMNS,
    INPUT_SHEET,
    OUTPUT_SHEETS,
    RESEARCH_COLUMNS,
    RESULT_COLUMNS,
    SUMMARY_COLUMNS,
    normalize_job_row,
)
from .utils import clean_text


def read_sheet(path):
    workbook = load_workbook(path)
    if INPUT_SHEET not in workbook.sheetnames:
        raise SystemExit(f"Input workbook must include a '{INPUT_SHEET}' sheet.")
    sheet = workbook[INPUT_SHEET]
    headers = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(values):
            continue
        row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        rows.append(normalize_job_row(row))
    return rows


SHEET_HEADER_COLORS = {
    "resumen_ejecucion": "385723",
    "vacantes_detectadas": "1F4E78",
    "preseleccionadas": "548235",
    "descartadas": "7F1D1D",
    "aplicadas": "7030A0",
    "requiere_intervencion": "C65911",
    "empresas_investigadas": "5F6368",
}


def write_rows(sheet, headers, rows):
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    style_sheet(sheet)


def style_sheet(sheet):
    color = SHEET_HEADER_COLORS.get(sheet.title, "1F4E78")
    header_fill = PatternFill("solid", fgColor=color)
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    sheet.freeze_panes = "A2"
    if sheet.max_column > 1:
        sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = get_column_letter(column[0].column)
        max_len = max(len(clean_text(cell.value)) for cell in column)
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)


def summary_rows(detected, shortlisted, discarded, applied, intervention_rows, research_rows):
    return [
        {"metrica": "fecha_generacion", "valor": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"metrica": "vacantes_detectadas", "valor": len(detected)},
        {"metrica": "preseleccionadas", "valor": len(shortlisted)},
        {"metrica": "descartadas", "valor": len(discarded)},
        {"metrica": "aplicadas", "valor": len(applied)},
        {"metrica": "requieren_intervencion", "valor": len(intervention_rows)},
        {"metrica": "empresas_investigadas", "valor": len(research_rows)},
        {"metrica": "cartas_generadas", "valor": sum(1 for row in shortlisted if clean_text(row.get("carta_de_interes_al_rol")))},
        {"metrica": "cache_hits", "valor": sum(1 for row in detected if clean_text(row.get("cache_hit")).lower() == "si")},
        {"metrica": "ignoradas_en_futuro", "valor": sum(1 for row in detected if clean_text(row.get("ignorar_en_futuro")).lower() == "si")},
    ]


def write_output(path, detected, shortlisted, discarded, applied, intervention_rows, research_rows):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in OUTPUT_SHEETS:
        workbook.create_sheet(name)
    write_rows(workbook["resumen_ejecucion"], SUMMARY_COLUMNS, summary_rows(detected, shortlisted, discarded, applied, intervention_rows, research_rows))
    write_rows(workbook["vacantes_detectadas"], RESULT_COLUMNS, detected)
    write_rows(workbook["preseleccionadas"], RESULT_COLUMNS, shortlisted)
    write_rows(workbook["descartadas"], RESULT_COLUMNS, discarded)
    write_rows(workbook["aplicadas"], RESULT_COLUMNS, applied)
    write_rows(workbook["requiere_intervencion"], RESULT_COLUMNS, intervention_rows)
    write_rows(workbook["empresas_investigadas"], RESEARCH_COLUMNS, research_rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    return save_workbook_with_fallback(workbook, path)


def save_workbook_with_fallback(workbook, path):
    try:
        workbook.save(path)
        return path
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = path.with_name(f"{path.stem}_{timestamp}{path.suffix}")
        workbook.save(fallback)
        return fallback


def create_template(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = INPUT_SHEET
    sheet.append(INPUT_COLUMNS)
    sheet.append([
        "Junior Full Stack Developer",
        "Acme SaaS",
        "LinkedIn",
        "https://example.com/job",
        "Buscamos developer con React, TypeScript, Node.js, APIs REST, PostgreSQL, Docker y CI/CD.",
        "Remoto Mexico",
        "remoto",
        "25000 MXN",
        date.today().isoformat(),
        "recruiter@example.com",
        "SaaS",
        "manual",
        "no",
        "pendiente",
        "no",
        "no",
        "",
        "",
        40,
        "Junior",
        "es",
        "https://example.com",
    ])
    style_sheet(sheet)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
